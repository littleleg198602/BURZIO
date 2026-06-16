"""Export cached analysis data to Excel with openpyxl."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .database import Database
from .phrase_extractor import extract_phrases

PERCENT_COLUMNS_BY_HEADER = {
    "Reakce +1 den",
    "Reakce +3 dny",
    "Reakce +5 dní",
    "Reakce +10 dní",
    "Průměrný výnos",
    "Úspěšnost růstu",
    "Průměrný pokles",
}


def export_excel(db: Database, tickers: tuple[str, ...], output_path: str) -> None:
    """Create a human-friendly workbook with instructions, ticker sheets, and phrase stats."""
    wb = Workbook()
    default = wb.active
    default.title = "NAVOD"
    _write_help_sheet(default)

    for ticker in tickers:
        sheet = wb.create_sheet(title=ticker[:31])
        _write_rows(sheet, _ticker_rows(db, ticker))
    phrase_sheet = wb.create_sheet(title="PHRASE_STATS")
    _write_rows(phrase_sheet, _phrase_rows(db))
    _add_empty_phrase_stats_note(phrase_sheet)
    wb.save(output_path)


def _write_help_sheet(sheet) -> None:
    sheet.append(["Yahoo News Impact Analyzer - jak číst výstup"])
    sheet.append([])
    sheet.append(["Listy s tickery", "Každý list NVDA/AAPL/MSFT/AMD/TSLA ukazuje zprávy, vytažené fráze, cenu při zprávě, budoucí ceny, reakce, směr a impact score."])
    sheet.append(["Reakce +1/+3/+5/+10", "Změna ceny od prvního obchodního dne po zprávě do daného obchodního okna. 0,025 znamená +2,5 %."])
    sheet.append(["Prázdná reakce", "Pro danou zprávu ještě není dost budoucích obchodních dní v cenových datech."])
    sheet.append(["PHRASE_STATS", "Souhrn frází z titulků/souhrnů. Vyšší skóre jistoty znamená častější a konzistentnější historický pohyb."])
    sheet.append(["Historie zpráv", "Ceny jsou 5 let dozadu. Zprávy se berou z Yahoo RSS a Yahoo Search v 5letém okně, ale Yahoo zdarma negarantuje kompletní archiv všech zpráv za 5 let."])
    sheet.append(["Když je PHRASE_STATS prázdný", "Nejde o chybu. Yahoo často vrací hlavně novější zprávy a pro +1/+3/+5/+10 obchodních dní ještě nemusí existovat budoucí ceny. Spusť analýzu později nebo nech minimum výskytů na 1."])
    sheet.append(["Důležité", "Výsledek je historická analýza, ne obchodní doporučení a ne důkaz příčiny."])
    sheet["A1"].font = Font(bold=True, size=14)
    for row in sheet.iter_rows(min_row=3, max_col=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110


def _ticker_rows(db: Database, ticker: str) -> list[list[object]]:
    news_by_guid = {row["guid"]: row for row in db.fetch_news(ticker)}
    grouped: dict[str, dict[int, dict[str, float | None]]] = {}
    for row in db.fetch_reactions(ticker):
        grouped.setdefault(row["guid"], {})[row["window_days"]] = {
            "return_pct": row["return_pct"],
            "start_price": row["start_price"],
            "end_price": row["end_price"],
        }
    rows: list[list[object]] = [[
        "Ticker",
        "Datum zprávy",
        "Titulek",
        "Shrnutí",
        "Fráze",
        "Cena při zprávě",
        "Cena +1 den",
        "Cena +3 dny",
        "Cena +5 dní",
        "Cena +10 dní",
        "Reakce +1 den",
        "Reakce +3 dny",
        "Reakce +5 dní",
        "Reakce +10 dní",
        "Směr +1 den",
        "Impact score",
        "Odkaz",
    ]]
    for guid, news in news_by_guid.items():
        reactions = grouped.get(guid, {})
        returns = [reactions.get(window, {}).get("return_pct") for window in (1, 3, 5, 10)]
        rows.append([
            ticker,
            news["published"],
            news["title"],
            news["summary"],
            ", ".join(sorted(extract_phrases(news["title"], news["summary"]))[:25]),
            _first_available_start_price(reactions),
            reactions.get(1, {}).get("end_price"),
            reactions.get(3, {}).get("end_price"),
            reactions.get(5, {}).get("end_price"),
            reactions.get(10, {}).get("end_price"),
            reactions.get(1, {}).get("return_pct"),
            reactions.get(3, {}).get("return_pct"),
            reactions.get(5, {}).get("return_pct"),
            reactions.get(10, {}).get("return_pct"),
            _direction(reactions.get(1, {}).get("return_pct")),
            _impact_score(returns),
            news["link"],
        ])
    return rows


def _first_available_start_price(reactions: dict[int, dict[str, float | None]]) -> float | None:
    for window in (1, 3, 5, 10):
        value = reactions.get(window, {}).get("start_price")
        if value is not None:
            return value
    return None


def _direction(return_pct: float | None) -> str:
    if return_pct is None:
        return "nelze spočítat"
    if return_pct > 0:
        return "nahoru"
    if return_pct < 0:
        return "dolů"
    return "beze změny"


def _impact_score(returns: list[float | None]) -> float | None:
    available = [abs(value) for value in returns if value is not None]
    if not available:
        return None
    return max(available) * 100


def _phrase_rows(db: Database) -> list[list[object]]:
    rows: list[list[object]] = [["Ticker", "Okno dní", "Fráze", "Počet výskytů", "Průměrný výnos", "Úspěšnost růstu", "Průměrný pokles", "Skóre jistoty"]]
    rows.extend([
        [row["ticker"], row["window_days"], row["phrase"], row["occurrences"], row["average_return"], row["win_rate_up"], row["average_down_move"], row["confidence_score"]]
        for row in db.fetch_phrase_stats()
    ])
    return rows


def _write_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)
    _format_table(sheet)


def _add_empty_phrase_stats_note(sheet) -> None:
    if sheet.title != "PHRASE_STATS" or sheet.max_row > 1:
        return
    sheet.append([
        "INFO",
        "",
        "Zatím není co spočítat: zprávy z Yahoo RSS jsou pravděpodobně moc nové, takže ještě nejsou dostupné budoucí ceny pro +1/+3/+5/+10 obchodních dní. Spusť analýzu později nebo zkontroluj list NAVOD.",
        "",
        "",
        "",
        "",
        "",
    ])
    for cell in sheet[2]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[2].height = 60


def _format_table(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    headers = [cell.value for cell in sheet[1]]
    percent_column_indexes = {index + 1 for index, header in enumerate(headers) if header in PERCENT_COLUMNS_BY_HEADER}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.column in percent_column_indexes and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        if header in {"Titulek", "Shrnutí", "Odkaz", "Fráze"}:
            width = {"Titulek": 58, "Shrnutí": 80, "Odkaz": 36, "Fráze": 60}[header]
        else:
            width = min(22, max(len(str(cell.value or "")) for cell in column_cells) + 2)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
